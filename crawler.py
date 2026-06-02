import os
import re
import time
import hashlib
import shutil
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse, urldefrag
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


HEADERS = {
    "User-Agent": "NSW Department PDF crawler - internal audit"
}

REQUEST_DELAY = 0.1
MAX_PAGES = 150
MAX_PDFS = 300


def normalise_url(url):
    url, _ = urldefrag(url)
    return url.rstrip("/")


def get_allowed_path_prefix(start_url):
    return urlparse(start_url).path.rstrip("/")


def get_allowed_domain(start_url):
    return urlparse(start_url).netloc


def is_allowed_page(url, allowed_domain, allowed_path_prefix):
    parsed = urlparse(url)
    return (
        parsed.netloc == allowed_domain
        and parsed.path.startswith(allowed_path_prefix)
    )


def looks_like_possible_file(url):
    path = urlparse(url).path.lower()
    return (
        path.endswith(".pdf")
        or "/media/" in path
        or "/document/" in path
        or "/file/" in path
        or "/sites/default/files/" in path
    )


def check_pdf_link(url):
    if url.lower().split("?")[0].endswith(".pdf"):
        try:
            response = requests.head(url, headers=HEADERS, timeout=15, allow_redirects=True)
            return True, response.status_code, response.url, response.headers.get("Content-Type", "")
        except Exception:
            return True, "", url, ""

    try:
        response = requests.head(url, headers=HEADERS, timeout=15, allow_redirects=True)
        content_type = response.headers.get("Content-Type", "").lower()

        if "application/pdf" in content_type:
            return True, response.status_code, response.url, content_type

        if response.status_code in [403, 405]:
            response = requests.get(
                url,
                headers=HEADERS,
                timeout=15,
                allow_redirects=True,
                stream=True
            )

            content_type = response.headers.get("Content-Type", "").lower()

            if "application/pdf" in content_type:
                return True, response.status_code, response.url, content_type

            return False, response.status_code, response.url, content_type

        return False, response.status_code, response.url, content_type

    except Exception as e:
        return False, "", url, f"error: {e}"


def clean_filename(name):
    return re.sub(r'[\\/*?:"<>|]', "_", name).strip()


def safe_filename(pdf_url):
    parsed = urlparse(pdf_url)
    name = os.path.basename(parsed.path)
    name = clean_filename(name)

    if not name or not name.lower().endswith(".pdf"):
        name = hashlib.md5(pdf_url.encode()).hexdigest() + ".pdf"

    return name


def get_unique_filepath(output_folder, filename):
    base, ext = os.path.splitext(filename)
    filepath = os.path.join(output_folder, filename)

    counter = 2
    while os.path.exists(filepath):
        filename = f"{base}_{counter}{ext}"
        filepath = os.path.join(output_folder, filename)
        counter += 1

    return filepath, filename


def download_pdf(pdf_url, output_folder):
    os.makedirs(output_folder, exist_ok=True)

    original_filename = safe_filename(pdf_url)
    filepath, final_filename = get_unique_filepath(output_folder, original_filename)

    response = requests.get(
        pdf_url,
        headers=HEADERS,
        timeout=60,
        allow_redirects=True,
        stream=True
    )
    response.raise_for_status()

    content_type = response.headers.get("Content-Type", "").lower()

    if "application/pdf" not in content_type and not response.url.lower().split("?")[0].endswith(".pdf"):
        return "", 0, response.status_code, response.url, f"skipped - not pdf, content-type: {content_type}"

    with open(filepath, "wb") as f:
        for chunk in response.iter_content(chunk_size=1024 * 1024):
            if chunk:
                f.write(chunk)

    return final_filename, os.path.getsize(filepath), response.status_code, response.url, "downloaded"


def format_file_size(size_bytes):
    if not size_bytes:
        return ""

    if size_bytes < 1024:
        return f"{size_bytes} B"

    if size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"

    return f"{size_bytes / (1024 * 1024):.2f} MB"


def get_page_title(soup):
    if soup.title and soup.title.string:
        return soup.title.string.strip()

    h1 = soup.find("h1")
    if h1:
        return h1.get_text(strip=True)

    return ""


def style_worksheet(ws):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def save_excel_log(pdf_records, crawled_pages, excel_path):
    wb = Workbook()

    # Sheet 1: PDF audit
    ws = wb.active
    ws.title = "PDF audit"

    pdf_headers = [
        "Source Page",
        "PDF URL",
        "Final Resolved PDF URL",
        "Downloaded Filename",
        "Status",
        "HTTP Status",
        "Duplicate?",
        "Original Row",
        "File Size",
        "File Size Bytes"
    ]

    ws.append(pdf_headers)

    for record in pdf_records:
        ws.append([record.get(header, "") for header in pdf_headers])

    for row in range(2, ws.max_row + 1):
        for col in [1, 2, 3]:
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"

    pdf_widths = {
        "A": 60,
        "B": 70,
        "C": 70,
        "D": 35,
        "E": 35,
        "F": 15,
        "G": 15,
        "H": 15,
        "I": 15,
        "J": 18
    }

    for col, width in pdf_widths.items():
        ws.column_dimensions[col].width = width

    style_worksheet(ws)

    # Sheet 2: Crawled URLs
    ws_pages = wb.create_sheet("Crawled URLs")

    page_headers = [
        "Page URL",
        "HTTP Status",
        "Page Title",
        "PDFs Found On Page"
    ]

    ws_pages.append(page_headers)

    for page in crawled_pages:
        ws_pages.append([
            page.get("Page URL", ""),
            page.get("HTTP Status", ""),
            page.get("Page Title", ""),
            page.get("PDFs Found On Page", 0)
        ])

    for row in range(2, ws_pages.max_row + 1):
        cell = ws_pages.cell(row=row, column=1)
        if cell.value:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"

    page_widths = {
        "A": 80,
        "B": 15,
        "C": 60,
        "D": 20
    }

    for col, width in page_widths.items():
        ws_pages.column_dimensions[col].width = width

    style_worksheet(ws_pages)

    wb.save(excel_path)


def crawl_site(start_url, job_folder):
    start_url = normalise_url(start_url)
    allowed_domain = get_allowed_domain(start_url)
    allowed_path_prefix = get_allowed_path_prefix(start_url)

    pdf_folder = os.path.join(job_folder, "downloaded_pdfs")
    excel_path = os.path.join(job_folder, "pdf_download_log.xlsx")

    zip_base = os.path.join(
        os.path.dirname(job_folder),
        "pdf_crawl_results_" + os.path.basename(job_folder)
    )

    visited_pages = set()
    pages_to_visit = [start_url]

    pdf_records = []
    crawled_pages = []

    seen_pdf_urls = set()
    pdf_first_seen_row = {}

    while pages_to_visit:
        if len(visited_pages) >= MAX_PAGES:
            break

        if len(seen_pdf_urls) >= MAX_PDFS:
            break

        current_url = pages_to_visit.pop(0)

        if current_url in visited_pages:
            continue

        if not is_allowed_page(current_url, allowed_domain, allowed_path_prefix):
            continue

        visited_pages.add(current_url)

        try:
            response = requests.get(
                current_url,
                headers=HEADERS,
                timeout=30,
                allow_redirects=True
            )
            response.raise_for_status()
        except Exception:
            crawled_pages.append({
                "Page URL": current_url,
                "HTTP Status": "failed",
                "Page Title": "",
                "PDFs Found On Page": 0
            })
            continue

        soup = BeautifulSoup(response.text, "html.parser")
        page_title = get_page_title(soup)
        pdfs_found_on_this_page = 0

        for link in soup.find_all("a", href=True):
            href = link["href"]
            absolute_url = normalise_url(urljoin(current_url, href))

            if looks_like_possible_file(absolute_url):
                is_pdf, check_status, resolved_url, content_type = check_pdf_link(absolute_url)

                if is_pdf:
                    pdfs_found_on_this_page += 1
                    is_duplicate = resolved_url in seen_pdf_urls

                    if is_duplicate:
                        original_row = pdf_first_seen_row.get(resolved_url, "")

                        pdf_records.append({
                            "Source Page": current_url,
                            "PDF URL": absolute_url,
                            "Final Resolved PDF URL": resolved_url,
                            "Downloaded Filename": "",
                            "Status": f"duplicate - original in row {original_row}",
                            "HTTP Status": check_status,
                            "Duplicate?": "Yes",
                            "Original Row": original_row,
                            "File Size": "",
                            "File Size Bytes": ""
                        })
                        continue

                    seen_pdf_urls.add(resolved_url)

                    try:
                        filename, file_size_bytes, http_status, final_url, status = download_pdf(
                            resolved_url,
                            pdf_folder
                        )
                        file_size = format_file_size(file_size_bytes)
                    except Exception as e:
                        filename = ""
                        file_size_bytes = ""
                        file_size = ""
                        http_status = check_status
                        final_url = resolved_url
                        status = f"failed: {e}"

                    pdf_records.append({
                        "Source Page": current_url,
                        "PDF URL": absolute_url,
                        "Final Resolved PDF URL": final_url,
                        "Downloaded Filename": filename,
                        "Status": status,
                        "HTTP Status": http_status,
                        "Duplicate?": "No",
                        "Original Row": "",
                        "File Size": file_size,
                        "File Size Bytes": file_size_bytes
                    })

                    excel_row_number = len(pdf_records) + 1
                    pdf_first_seen_row[resolved_url] = excel_row_number

            elif is_allowed_page(absolute_url, allowed_domain, allowed_path_prefix):
                if absolute_url not in visited_pages and absolute_url not in pages_to_visit:
                    pages_to_visit.append(absolute_url)

        crawled_pages.append({
            "Page URL": current_url,
            "HTTP Status": response.status_code,
            "Page Title": page_title,
            "PDFs Found On Page": pdfs_found_on_this_page
        })

        time.sleep(REQUEST_DELAY)

    save_excel_log(pdf_records, crawled_pages, excel_path)

    zip_path = shutil.make_archive(
        zip_base,
        "zip",
        job_folder
    )

    return zip_path
